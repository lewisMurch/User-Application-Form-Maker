import openpyxl

# Load the Excel file
file_path = 'example_excel_file.xlsx'
wb = openpyxl.load_workbook(file_path)
sheet = wb.active


col = column=sheet.max_column +1
# Iterate through rows and update the 'CRSID' column
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=3):
    cell = row[0]
    if cell.value:
        # Extract text before the '@' sign
        text = cell.value.split('@')[0]
        # Create a new 'CRSID' column and set the value
        sheet.cell(row=cell.row, column=col, value=text)

# Save the modified Excel file
wb.save('new_excel_file.xlsx')
